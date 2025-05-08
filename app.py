import streamlit as st
import openpyxl
import requests
import os
import re
import time
from collections import defaultdict
import concurrent.futures

# Load configuration from Streamlit Secrets
def load_config():
    config = st.secrets["openwebui"]
    return {
        "endpoint": config["endpoint"],
        "api_key": config["api_key"]
    }

config = load_config()
api_endpoint = config['endpoint']
api_key = config['api_key']

# Load translation prompt template
def load_prompt_template():
    try:
        with open('prompt.txt', 'r', encoding='utf-8') as f:
            return f.read().strip()
    except FileNotFoundError:
        st.error("prompt.txt file missing! Create it with the translation template.")
        st.stop()

PROMPT_TEMPLATE = load_prompt_template()

# Language style mapping
def get_language_style_for_class(student_class):
    class_to_age_mapping = {
        'A': 'for a 3-year-old child',
        'B': 'for a 4-year-old child',
        'C': 'for a 5-year-old child',
        '1': 'for a 6-year-old child',
        '2': 'for a 7-year-old child',
        '3': 'for a 8-year-old child',
        '4': 'for a 9-year-old child',
        '5': 'for a 10-year-old child'
    }
    return class_to_age_mapping.get(student_class, 'for a 6-year old')

# Translation cache to avoid redundant API calls
translation_cache = {}

# Improved text processing utilities
def is_hindi_text(text):
    # More robust Hindi text detection - requires at least 30% Hindi characters for mixed text
    if not text or not isinstance(text, str):
        return False
    
    text = text.strip()
    if not text:
        return False
        
    hindi_chars = len(re.findall(r'[\u0900-\u097F]', text))
    total_chars = len(text)
    
    # Either has multiple Hindi characters and makes up significant portion of text
    # or is predominantly Hindi
    return (hindi_chars >= 2 and (hindi_chars / total_chars) > 0.3) or hindi_chars > 5

def split_text_parts(text):
    # More comprehensive splitting to handle various delimiters
    if not text:
        return []
    # Split by HTML tags, brackets, and other common separators
    return re.split(r'(<br>|<[^>]*>|\[.*?\]|\(.*?\))', text)

# API request handler with improved error handling and exponential backoff
def openwebui_request(cleaned_text, language_style, model="us.deepseek.r1-v1:0"):
    # Check cache first
    cache_key = f"{cleaned_text}_{language_style}"
    if cache_key in translation_cache:
        return translation_cache[cache_key]
    
    try:
        prompt = PROMPT_TEMPLATE.format(
            language_style=language_style,
            text=cleaned_text
        )
        
        payload = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.3
        }
        headers = {"Authorization": f"Bearer {api_key}"}

        max_attempts = 5
        for attempt in range(max_attempts):
            try:
                # Exponential backoff
                backoff_time = min(2 ** attempt, 16)
                
                response = requests.post(
                    api_endpoint,
                    json=payload,
                    headers=headers,
                    timeout=60  # Increased timeout
                )
                
                if response.status_code == 200:
                    result = response.json().get("choices", [{}])[0].get("message", {}).get("content", cleaned_text)
                    # Store in cache
                    translation_cache[cache_key] = result
                    return result
                elif response.status_code == 429:  # Rate limit
                    st.warning(f"Rate limit hit. Waiting {backoff_time} seconds...")
                    time.sleep(backoff_time)
                    continue
                else:
                    st.warning(f"API returned status code {response.status_code}. Retrying...")
                    time.sleep(backoff_time)
            except requests.exceptions.RequestException as e:
                st.warning(f"Attempt {attempt+1}/{max_attempts} failed: {str(e)}")
                time.sleep(backoff_time)
                
        st.error(f"Translation failed after {max_attempts} attempts")
        return cleaned_text

    except Exception as e:
        st.error(f"API Error: {str(e)}")
        return cleaned_text

# Batch translation function
def batch_translate(texts_with_styles, max_batch_size=5, model="us.deepseek.r1-v1:0"):
    results = {}
    
    # Group by language style to optimize caching
    style_groups = defaultdict(list)
    for text, style in texts_with_styles:
        style_groups[style].append(text)
    
    for style, texts in style_groups.items():
        # Process in smaller batches
        for i in range(0, len(texts), max_batch_size):
            batch = texts[i:i+max_batch_size]
            
            # Process batch in parallel
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_batch_size) as executor:
                future_to_text = {
                    executor.submit(openwebui_request, text, style, model): text 
                    for text in batch
                }
                
                for future in concurrent.futures.as_completed(future_to_text):
                    text = future_to_text[future]
                    try:
                        results[text] = future.result()
                    except Exception as e:
                        st.error(f"Error translating '{text}': {str(e)}")
                        results[text] = text
    
    return results

# Excel processing logic with batch processing and progress tracking
def process_excel(file, model="us.deepseek.r1-v1:0", batch_size=10):
    try:
        # Show loading spinner while processing
        with st.spinner("Loading Excel file..."):
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

        class_column_index = None
        valid_classes = ['A','B','C','1','2','3','4','5']

        # Find the Class column
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value == 'Class':
                class_column_index = col[0].col_idx - 1
                break

        if class_column_index is None:
            st.error("'Class' column not found.")
            return None, []

        # Count total cells for progress tracking
        total_rows = sheet.max_row - 1  # Excluding header
        st.info(f"Processing {total_rows} rows...")
        
        # Create progress bar
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        output_data = []
        error_count = 0
        
        # First pass: collect all Hindi text segments with their language styles
        status_text.text("Phase 1/2: Analyzing text...")
        texts_to_translate = []
        cell_mapping = {}  # Maps (row_idx, col_idx) to list of (part_idx, is_hindi)
        
        # Process in chunks of rows for better UI responsiveness
        chunk_size = 10
        num_chunks = (total_rows + chunk_size - 1) // chunk_size
        
        for chunk_idx in range(num_chunks):
            start_row = 2 + chunk_idx * chunk_size
            end_row = min(start_row + chunk_size, sheet.max_row + 1)
            
            for row_idx in range(start_row, end_row):
                try:
                    row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
                    
                    # Update progress
                    progress = int((row_idx - 1) / total_rows * 50)  # First phase is 50% of progress
                    progress_bar.progress(progress)
                    
                    # Get student class and language style
                    student_class = str(row[class_column_index].value).strip() if row[class_column_index].value else ""
                    if student_class not in valid_classes:
                        st.warning(f"Row {row_idx}: Invalid class '{student_class}'")
                        error_count += 1
                        continue

                    language_style = get_language_style_for_class(student_class)

                    # Process each cell in the row
                    for col_idx, cell in enumerate(row):
                        if isinstance(cell.value, str) and cell.value.strip():
                            parts = split_text_parts(cell.value)
                            cell_parts = []
                            
                            for part_idx, part in enumerate(parts):
                                if is_hindi_text(part):
                                    cleaned = part.strip()
                                    if cleaned:
                                        texts_to_translate.append((cleaned, language_style))
                                        cell_parts.append((part_idx, True))  # True means Hindi text
                                else:
                                    cell_parts.append((part_idx, False))  # False means non-Hindi text
                            
                            # Store mapping for reconstruction
                            if cell_parts:
                                cell_mapping[(row_idx, col_idx)] = {
                                    'parts': parts,
                                    'hindi_indices': cell_parts,
                                    'language_style': language_style
                                }

                    if error_count > 5:
                        st.error("Too many errors - stopping processing")
                        break

                except Exception as e:
                    st.error(f"Error analyzing row {row_idx}: {str(e)}")
                    continue
        
        # Second pass: batch translate all collected Hindi text
        status_text.text(f"Phase 2/2: Translating {len(texts_to_translate)} text segments...")
        
        # Remove duplicates while preserving order
        unique_texts = []
        seen = set()
        for text, style in texts_to_translate:
            if text not in seen:
                unique_texts.append((text, style))
                seen.add(text)
        
        # Batch translate using the provided batch size
        translations = {}
        if unique_texts:
            # batch_size is already set from the function parameter
            total_batches = (len(unique_texts) + batch_size - 1) // batch_size
            
            for i in range(total_batches):
                batch_start = i * batch_size
                batch_end = min((i + 1) * batch_size, len(unique_texts))
                current_batch = unique_texts[batch_start:batch_end]
                
                # Update progress
                progress = 50 + int(i / total_batches * 40)  # Second phase is 40% of progress
                progress_bar.progress(progress)
                status_text.text(f"Translating batch {i+1}/{total_batches}...")
                
                # Get translations for this batch using the specified model
                batch_results = batch_translate(current_batch, max_batch_size=batch_size, model=model)
                translations.update({text: trans for text, trans in batch_results.items()})
        
        # Third pass: reconstruct cells with translations
        status_text.text("Updating Excel with translations...")
        
        for (row_idx, col_idx), cell_data in cell_mapping.items():
            try:
                parts = cell_data['parts']
                hindi_indices = cell_data['hindi_indices']
                language_style = cell_data['language_style']
                
                # Reconstruct the cell with translations
                translated_parts = []
                for part_idx, is_hindi in hindi_indices:
                    if is_hindi:
                        cleaned = parts[part_idx].strip()
                        translated = translations.get(cleaned, cleaned)
                        translated_parts.append(translated)
                    else:
                        translated_parts.append(parts[part_idx])
                
                # Update the cell value
                sheet.cell(row=row_idx, column=col_idx+1).value = ''.join(translated_parts)
                
                # Add to output data for display
                output_data.append({
                    "Original": ''.join(parts),
                    "Translated": ''.join(translated_parts)
                })
                
            except Exception as e:
                st.error(f"Error updating cell at row {row_idx}, column {col_idx}: {str(e)}")
        
        # Complete progress bar
        progress_bar.progress(100)
        status_text.text("Processing complete!")
        
        # Display translation statistics
        st.success(f"Translated {len(translations)} unique Hindi text segments across {len(cell_mapping)} cells.")
        
        return workbook, output_data

    except Exception as e:
        st.error(f"Excel processing error: {str(e)}")
        return None, []

# Streamlit UI with improved user experience
def main():
    st.set_page_config(
        page_title="Hindi-English Excel Translator",
        page_icon="🔤",
        layout="wide"
    )
    
    st.title('Hindi-English Excel Translator')
    
    with st.expander("ℹ️ Instructions & Tips", expanded=True):
        st.markdown("""
        ### Instructions
        - Upload an Excel file with a 'Class' column
        - Supported classes: A, B, C, 1-5 (each maps to age-appropriate language)
        - Hindi text will be automatically detected and translated
        
        ### Performance Tips
        - Translation is done in batches for better performance
        - Duplicate text is only translated once (cached)
        - Progress indicators show real-time status
        """)
    
    # Create two columns for settings and file upload
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.subheader("Settings")
        model = st.selectbox(
            "Translation Model",
            ["us.deepseek.r1-v1:0", "us.deepseek.r1-v1:1", "us.deepseek.r1-v1:2"],
            index=0,
            help="Select the model to use for translation"
        )
        
        batch_size = st.slider(
            "Batch Size",
            min_value=1,
            max_value=20,
            value=10,
            help="Number of text segments to translate in parallel (higher values may be faster but could hit rate limits)"
        )
        
        st.markdown("---")
        
        # Cache statistics
        if translation_cache:
            st.success(f"✅ Cache contains {len(translation_cache)} translations")
            if st.button("Clear Cache"):
                translation_cache.clear()
                st.success("Cache cleared!")
    
    with col2:
        st.subheader("Upload File")
        uploaded_file = st.file_uploader("Choose Excel file", type="xlsx")
        
        if uploaded_file:
            # File info
            file_details = {
                "Filename": uploaded_file.name,
                "File size": f"{uploaded_file.size / 1024:.1f} KB"
            }
            st.json(file_details)
            
            # Process button
            if st.button("Start Translation", type="primary"):
                start_time = time.time()
                
                # Process the file with the selected model and batch size
                workbook, output_data = process_excel(
                    uploaded_file,
                    model=model,
                    batch_size=batch_size
                )
                
                if workbook:
                    # Calculate processing time
                    processing_time = time.time() - start_time
                    
                    # Success message with stats
                    st.success(f"✅ Translation complete in {processing_time:.1f} seconds!")
                    
                    # Save and provide download
                    output_filename = f"{os.path.splitext(uploaded_file.name)[0]}_en.xlsx"
                    
                    with st.spinner("Saving file..."):
                        workbook.save(output_filename)
                        with open(output_filename, "rb") as f:
                            st.download_button(
                                "📥 Download Translated File",
                                f,
                                file_name=output_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    
                    # Show sample translations
                    if output_data:
                        with st.expander("View Sample Translations", expanded=False):
                            st.write("Sample of translations (up to 5):")
                            for i, item in enumerate(output_data[:5]):
                                st.markdown(f"**Sample {i+1}:**")
                                st.markdown(f"Original: {item['Original']}")
                                st.markdown(f"Translated: {item['Translated']}")
                                st.markdown("---")

if __name__ == '__main__':
    main()